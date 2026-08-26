#!/usr/bin/env python3
"""
Batch Gmail draft creator — Google Doc (message) + Google Sheet (recipients).

Reads the email copy from a Google Doc, the recipient list from a Google
Sheet, merges the two, and creates a DRAFT in your Gmail for each row.
It never sends anything.

Self-contained: the only dependencies are the Google API client libraries.

    python batch_emailer.py --doc <doc-url> --sheet <sheet-url> [--dry-run]

── The Doc ──────────────────────────────────────────────────────────────
First line may set the subject; everything after it is the body:

    Subject: Quick intro — {{first_name}} <> AI Fund

    Hi {{first_name}},

    Saw {{fund_name}} led the round in ...

Bold, italics, links, and bullet lists are preserved (the draft is sent as
HTML with a plain-text alternative). Tables and inline images are skipped.
{{placeholders}} match Sheet column headers, case/space insensitive, so a
column "Fund Name" fills {{fund_name}} or {{Fund Name}}.

── The Sheet ────────────────────────────────────────────────────────────
Row 1 is the header row. Required: an "email" column (or "email address" /
"recipient"). Everything else is optional:

    cc, bcc          extra addresses, comma or semicolon separated
    subject          per-row subject, overrides the Doc's
    attachments      local file path(s), semicolon separated
    status           written back as "Draft created <date>" (skipped on re-run)
    draft_id         written back with the Gmail draft id

Rows with a non-empty status are skipped unless --force, so re-running after
a partial failure won't double-draft.

── Setup (one time) ─────────────────────────────────────────────────────
Needs credentials.json — an OAuth "Desktop app" client — in this directory,
and the Gmail, Google Docs and Google Sheets APIs enabled in that Cloud
project. The first run opens a browser and writes token.json here; after
that it runs unattended. Full walkthrough in README.md.
"""
from __future__ import annotations

import argparse
import csv
import html as html_lib
import mimetypes
import os
import re
import sys
from datetime import date
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from typing import Optional

import base64
import warnings

# The Google libs warn on every import that Python 3.9 is past EOL. Nothing to
# act on here — it just buries the run output.
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=".*OpenSSL.*")

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
except ImportError:
    sys.exit(
        "Missing Google libraries. Run:\n"
        "  pip install google-auth google-auth-oauthlib google-api-python-client"
    )


# ============================================================
# CONFIG
# ============================================================
# Resolved next to this script, so it works from any working directory.
HERE = Path(__file__).resolve().parent
CREDENTIALS_FILE = str(HERE / "credentials.json")
TOKEN_FILE = str(HERE / "token.json")
LOG_PATH = str(HERE / "drafts_log.csv")

SCOPES = [
    "https://www.googleapis.com/auth/gmail.compose",
    "https://www.googleapis.com/auth/documents.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
    # Needed to detect the file type, and to download recipient lists that are
    # uploaded .xlsx files rather than native Google Sheets.
    "https://www.googleapis.com/auth/drive.readonly",
]

NATIVE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Header names accepted for the recipient address, in priority order.
EMAIL_COLUMNS = ["email", "email_address", "recipient", "recipient_email", "to"]

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# {{field}} — strict: an unknown field is an error (the row gets skipped).
PLACEHOLDER_RE = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")
# [field] — lenient: filled only when it names a real Sheet column, so ordinary
# bracketed prose ("[Luma link]", "[TBD]") is left untouched.
BRACKET_RE = re.compile(r"\[\s*([A-Za-z][A-Za-z0-9 _-]{0,40})\s*\]")

# A line like "Email 2" or "Email 2: TBD" — used to find where a --section ends.
SECTION_HEAD_RE = re.compile(r"^([A-Za-z][A-Za-z ]{0,20}?)\s*#?\s*(\d+)\b")
# A divider line: ---, ———, —------, ___
DIVIDER_RE = re.compile(r"^[\s\-—–_=*]{3,}$")

# Wrapper for the HTML body — matches the look of the existing outreach emails.
HTML_WRAPPER = (
    '<div style="font-family: Arial, sans-serif; font-size: 14px; '
    'color: #1a1a1a; line-height: 1.6;">\n{body}\n</div>'
)


# ============================================================
# HELPERS
# ============================================================
def norm_key(s: str) -> str:
    """'Fund Name' / 'fund-name' / 'Fund  name' → 'fund_name'."""
    return re.sub(r"[^a-z0-9]+", "_", s.strip().lower()).strip("_")


def extract_id(url_or_id: str, kind: str) -> str:
    """Accept a full Docs/Sheets URL or a bare file id."""
    s = url_or_id.strip()
    m = re.search(r"/d/([a-zA-Z0-9_-]{20,})", s)
    if m:
        return m.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", s):
        return s
    sys.exit(f"❌ Could not read a {kind} id from: {url_or_id}")


def split_addresses(value: str) -> list:
    if not value:
        return []
    parts = re.split(r"[,;]", value)
    return [p.strip() for p in parts if p.strip()]


# ============================================================
# AUTH
# ============================================================
def authenticate(port: int = 0):
    if not os.path.exists(CREDENTIALS_FILE):
        sys.exit(
            f"❌ credentials.json not found in {HERE}.\n"
            "   Create an OAuth 'Desktop app' client in the Google Cloud console,\n"
            "   download the JSON, and save it there. See README.md."
        )

    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds:
            print("  🌐 Opening browser for Google auth (one-time)...\n")
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            try:
                creds = flow.run_local_server(port=port)
            except Exception as e:
                sys.exit(f"❌ Google auth failed: {e}")
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())

    gmail = build("gmail", "v1", credentials=creds)
    docs = build("docs", "v1", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)
    drive = build("drive", "v3", credentials=creds)

    profile = gmail.users().getProfile(userId="me").execute()
    sender = profile.get("emailAddress", "")
    print(f"  ✅ Signed in as {sender}")
    return gmail, docs, sheets, drive, sender


# ============================================================
# GOOGLE DOC → subject + html/text body
# ============================================================
def _run_text(run: dict) -> str:
    """Text of one run. Docs encodes a soft line break (shift+enter) as \\x0b."""
    text = run.get("content", "")
    if text.endswith("\n"):
        text = text[:-1]
    return text.replace("\x0b", "\n")


def _run_html(run: dict) -> str:
    text = _run_text(run)
    if not text:
        return ""
    style = run.get("textStyle", {})
    out = html_lib.escape(text).replace("\n", "<br>")
    if style.get("bold"):
        out = f"<b>{out}</b>"
    if style.get("italic"):
        out = f"<i>{out}</i>"
    if style.get("underline") and not style.get("link"):
        out = f"<u>{out}</u>"
    link = (style.get("link") or {}).get("url")
    if link:
        out = f'<a href="{html_lib.escape(link, quote=True)}">{out}</a>'
    return out


def _paragraph_blocks(doc: dict) -> list:
    """Flatten the Doc into blocks: {'kind','text','html','level','ordered'}."""
    lists = doc.get("lists", {})
    blocks = []
    skipped_tables = 0

    for element in doc.get("body", {}).get("content", []):
        if "table" in element:
            skipped_tables += 1
            continue
        para = element.get("paragraph")
        if not para:
            continue

        text_parts, html_parts = [], []
        for el in para.get("elements", []):
            run = el.get("textRun")
            if not run:
                continue
            text_parts.append(_run_text(run))
            html_parts.append(_run_html(run))

        text = "".join(text_parts).rstrip()
        inner_html = "".join(html_parts)

        bullet = para.get("bullet")
        if bullet:
            level = bullet.get("nestingLevel", 0)
            list_id = bullet.get("listId")
            glyph = ""
            try:
                glyph = (
                    lists[list_id]["listProperties"]["nestingLevels"][level]
                    .get("glyphType", "")
                )
            except (KeyError, IndexError):
                pass
            blocks.append({
                "kind": "li",
                "text": text,
                "html": inner_html,
                "level": level,
                "ordered": bool(glyph) and glyph != "GLYPH_TYPE_UNSPECIFIED",
            })
            continue

        style = (para.get("paragraphStyle") or {}).get("namedStyleType", "")
        kind = "h" if style.startswith("HEADING") else "p"
        blocks.append({
            "kind": kind,
            "text": text,
            "html": inner_html,
            "level": 0,
            "ordered": False,
        })

    if skipped_tables:
        print(f"  ⚠️  Skipped {skipped_tables} table(s) in the Doc — not supported.")
    return blocks


def _render(blocks: list) -> tuple:
    """Blocks → (plain_text, html)."""
    text_lines, html_parts = [], []
    open_list = None  # 'ul' | 'ol'

    def close_list():
        nonlocal open_list
        if open_list:
            html_parts.append(f"</{open_list}>")
            open_list = None

    for b in blocks:
        if b["kind"] == "li":
            tag = "ol" if b["ordered"] else "ul"
            if open_list != tag:
                close_list()
                html_parts.append(f"<{tag}>")
                open_list = tag
            html_parts.append(f"<li>{b['html']}</li>")
            text_lines.append(("  " * b["level"]) + "• " + b["text"])
            continue

        close_list()
        if not b["text"].strip():
            text_lines.append("")
            continue
        if b["kind"] == "h":
            html_parts.append(f"<p><b>{b['html']}</b></p>")
        else:
            html_parts.append(f"<p>{b['html']}</p>")
        text_lines.append(b["text"])

    close_list()

    # Trim leading/trailing blank lines from the plain-text version.
    while text_lines and not text_lines[0].strip():
        text_lines.pop(0)
    while text_lines and not text_lines[-1].strip():
        text_lines.pop()

    return "\n".join(text_lines), HTML_WRAPPER.format(body="\n".join(html_parts))


def _slice_section(blocks: list, section: str) -> list:
    """Keep only the blocks belonging to one section of a multi-draft Doc.

    The section starts at the line naming it ("Email 1") and ends at the next
    line of the same shape ("Email 2", "Email 2: TBD") or at a divider line.
    """
    want = norm_key(section)
    start = None
    for i, b in enumerate(blocks):
        if norm_key(b["text"]).startswith(want):
            start = i
            break
    if start is None:
        headings = [
            b["text"].strip() for b in blocks
            if SECTION_HEAD_RE.match(b["text"].strip())
        ]
        sys.exit(
            f"❌ No section matching '{section}' in the Doc."
            + (f" Sections found: {', '.join(headings)}" if headings else "")
        )

    end = len(blocks)
    for j in range(start + 1, len(blocks)):
        t = blocks[j]["text"].strip()
        if not t:
            continue
        if SECTION_HEAD_RE.match(t) or DIVIDER_RE.match(t):
            end = j
            break

    print(f"  ✂️  Section '{blocks[start]['text'].strip()}' "
          f"({end - start - 1} lines)")
    return blocks[start + 1:end]


def load_doc(docs_service, doc_id: str, section: Optional[str] = None) -> tuple:
    """Returns (subject_or_None, body_text, body_html)."""
    doc = docs_service.documents().get(documentId=doc_id).execute()
    print(f"  📄 Doc: {doc.get('title', '(untitled)')}")
    blocks = _paragraph_blocks(doc)

    if section:
        blocks = _slice_section(blocks, section)

    subject = None
    for i, b in enumerate(blocks):
        if not b["text"].strip():
            continue
        m = re.match(r"^subject\s*:\s*(.+)$", b["text"].strip(), re.IGNORECASE)
        if m:
            subject = m.group(1).strip()
            blocks.pop(i)
        break  # only the first non-empty block can be the subject line

    body_text, body_html = _render(blocks)
    return subject, body_text, body_html


# ============================================================
# GOOGLE SHEET → recipient rows
# ============================================================
def _match_tab(tab: Optional[str], tab_names: list) -> str:
    """Exact, then case-insensitive, then punctuation-insensitive match."""
    if not tab:
        return tab_names[0]
    if tab in tab_names:
        return tab
    for t in tab_names:
        if t.lower() == tab.lower():
            return t
    for t in tab_names:
        if norm_key(t) == norm_key(tab):
            return t
    # Last resort: ignore all punctuation, so "CEOs & cofounders" finds
    # "CEOs & Co-founders".
    squash = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())
    for t in tab_names:
        if squash(t) == squash(tab):
            return t
    sys.exit(f"❌ Tab '{tab}' not found. Tabs: {', '.join(tab_names)}")


def _find_header_row(grid: list, header_row: Optional[int]) -> int:
    """Index of the header row — the first row that names an email column.

    Real-world sheets often carry a title and a legend above the real header.
    """
    if header_row:
        return header_row - 1  # 1-indexed on the CLI
    for i, raw in enumerate(grid[:25]):
        keys = [norm_key(str(c)) for c in raw if str(c).strip()]
        if len(keys) >= 2 and any(k in EMAIL_COLUMNS for k in keys):
            return i
    sys.exit(
        "❌ Couldn't find a header row with an email column in the first 25 rows.\n"
        "   Pass --header-row N (1-indexed) to point at it."
    )


def _grid_to_rows(grid: list, header_idx: int) -> tuple:
    headers = [norm_key(str(c)) for c in grid[header_idx]]
    rows = []
    for offset, raw in enumerate(grid[header_idx + 1:]):
        row_number = header_idx + 2 + offset  # 1-indexed row in the file
        raw = list(raw) + [""] * (len(headers) - len(raw))
        row = {h: str(raw[i]).strip() for i, h in enumerate(headers) if h}
        row["_row_number"] = row_number
        if any(v for k, v in row.items() if k != "_row_number"):
            rows.append(row)
    return headers, rows


def load_sheet(sheets_service, drive_service, file_id: str, tab: Optional[str],
               header_row: Optional[int] = None) -> tuple:
    """Returns (headers, rows, tab_name, writable).

    Handles both native Google Sheets and .xlsx files uploaded to Drive. An
    .xlsx can be read but not written back to, so `writable` is False there.
    """
    meta = drive_service.files().get(
        fileId=file_id, fields="name,mimeType", supportsAllDrives=True
    ).execute()
    name, mime = meta.get("name", "?"), meta.get("mimeType", "")

    if mime == NATIVE_SHEET_MIME:
        info = sheets_service.spreadsheets().get(spreadsheetId=file_id).execute()
        tab_names = [s["properties"]["title"] for s in info.get("sheets", [])]
        tab_name = _match_tab(tab, tab_names)
        print(f"  📊 Sheet: {name} → tab '{tab_name}'")
        grid = (
            sheets_service.spreadsheets().values()
            .get(spreadsheetId=file_id, range=f"'{tab_name}'")
            .execute().get("values", [])
        )
        writable = True

    elif mime == XLSX_MIME:
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit(
                "❌ This recipient list is an uploaded .xlsx file. Reading it needs\n"
                "   openpyxl:  .venv/bin/pip install openpyxl"
            )
        import io
        data = drive_service.files().get_media(
            fileId=file_id, supportsAllDrives=True
        ).execute()
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        tab_name = _match_tab(tab, wb.sheetnames)
        print(f"  📊 Workbook (.xlsx): {name} → tab '{tab_name}'")
        grid = [
            ["" if c is None else c for c in r]
            for r in wb[tab_name].iter_rows(values_only=True)
        ]
        wb.close()
        writable = False

    else:
        sys.exit(
            f"❌ '{name}' is not a spreadsheet ({mime}).\n"
            "   Point --sheet at a Google Sheet or an uploaded .xlsx file."
        )

    if not grid:
        sys.exit("❌ That tab is empty.")

    header_idx = _find_header_row(grid, header_row)
    if header_idx:
        print(f"  ↳ header row is row {header_idx + 1}")
    headers, rows = _grid_to_rows(grid, header_idx)
    return headers, rows, tab_name, writable


def pick_email(row: dict) -> str:
    for col in EMAIL_COLUMNS:
        if row.get(col):
            return row[col].strip()
    return ""


def enrich(row: dict) -> dict:
    """Derive first_name/last_name from a name column when not given directly."""
    if not row.get("first_name"):
        full = row.get("name") or row.get("full_name") or row.get("contact_name") or ""
        parts = full.split()
        if parts:
            row["first_name"] = parts[0]
            if not row.get("last_name") and len(parts) > 1:
                row["last_name"] = parts[-1]
    return row


# ============================================================
# MERGE
# ============================================================
def parse_maps(specs: list) -> dict:
    """['name=first_name'] → {'name': 'first_name'} — placeholder → column."""
    out = {}
    for spec in specs:
        if "=" not in spec:
            sys.exit(f"❌ --map needs PLACEHOLDER=COLUMN, got '{spec}'")
        placeholder, column = spec.split("=", 1)
        out[norm_key(placeholder)] = norm_key(column)
    return out


def merge(template: str, row: dict, escape: bool, aliases: Optional[dict] = None) -> tuple:
    """Substitute {{placeholders}} and [placeholders].

    {{field}} is strict — an unknown field is reported as missing.
    [field] is lenient — filled only when it names a real column, so prose like
    "[Luma link]" survives untouched.

    Returns (result, list_of_missing_fields).
    """
    missing = []
    aliases = aliases or {}

    def resolve(raw_key):
        key = norm_key(raw_key)
        return aliases.get(key, key)

    def value_of(key):
        v = row[key]
        return html_lib.escape(v) if escape else v

    def sub_strict(m):
        key = resolve(m.group(1))
        if row.get(key):
            return value_of(key)
        missing.append(m.group(1).strip())
        return m.group(0)

    def sub_lenient(m):
        key = resolve(m.group(1))
        return value_of(key) if row.get(key) else m.group(0)

    out = PLACEHOLDER_RE.sub(sub_strict, template)
    out = BRACKET_RE.sub(sub_lenient, out)
    return out, missing


def parse_filters(specs: list) -> list:
    """['tier=1', 'status=active'] → [('tier', {'1'}), ('status', {'active'})]."""
    out = []
    for spec in specs:
        if "=" not in spec:
            sys.exit(f"❌ --filter needs COLUMN=VALUE, got '{spec}'")
        col, values = spec.split("=", 1)
        wanted = {v.strip().lower() for v in values.split(",") if v.strip()}
        if not wanted:
            sys.exit(f"❌ --filter '{spec}' has no value")
        out.append((norm_key(col), wanted))
    return out


def _cell(row: dict, col: str) -> str:
    cell = str(row.get(col, "")).strip().lower()
    # "1" should match a spreadsheet's numeric 1 that arrives as "1.0"
    return cell[:-2] if cell.endswith(".0") else cell


def passes_filters(row: dict, filters: list) -> bool:
    return all(_cell(row, col) in wanted for col, wanted in filters)


def is_excluded(row: dict, excludes: list) -> bool:
    return any(_cell(row, col) in unwanted for col, unwanted in excludes)


# ============================================================
# DRAFT CREATION
# ============================================================
def build_message(sender, to, subject, body_text, body_html, cc, bcc, attachments):
    inner = MIMEMultipart("alternative")
    inner.attach(MIMEText(body_text, "plain", "utf-8"))
    inner.attach(MIMEText(body_html, "html", "utf-8"))

    if attachments:
        msg = MIMEMultipart("mixed")
        msg.attach(inner)
        for path in attachments:
            p = Path(path).expanduser()
            if not p.is_file():
                print(f"      ⚠️  Attachment not found, skipping: {p}")
                continue
            ctype, encoding = mimetypes.guess_type(str(p))
            if ctype is None or encoding is not None:
                ctype = "application/octet-stream"
            maintype, subtype = ctype.split("/", 1)
            part = MIMEBase(maintype, subtype)
            part.set_payload(p.read_bytes())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=p.name)
            msg.attach(part)
    else:
        msg = inner

    msg["To"] = to
    msg["From"] = sender
    msg["Subject"] = subject
    if cc:
        msg["Cc"] = ", ".join(cc)
    if bcc:
        msg["Bcc"] = ", ".join(bcc)
    return msg


def create_draft(gmail_service, message) -> str:
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode("utf-8")
    draft = (
        gmail_service.users()
        .drafts()
        .create(userId="me", body={"message": {"raw": raw}})
        .execute()
    )
    return draft.get("id", "")


# ============================================================
# SHEET WRITE-BACK
# ============================================================
def write_back(sheets_service, sheet_id, tab_name, headers, updates):
    """updates: list of (row_number, column_key, value)."""
    if not updates:
        return
    data = []
    for row_number, col_key, value in updates:
        if col_key not in headers:
            continue
        col_idx = headers.index(col_key)
        a1 = _col_letter(col_idx + 1)
        data.append({
            "range": f"'{tab_name}'!{a1}{row_number}",
            "values": [[value]],
        })
    if not data:
        return
    sheets_service.spreadsheets().values().batchUpdate(
        spreadsheetId=sheet_id,
        body={"valueInputOption": "RAW", "data": data},
    ).execute()
    print(f"  ✍️  Updated {len(data)} cell(s) in the sheet")


def _col_letter(n: int) -> str:
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


# ============================================================
# MAIN
# ============================================================
def main():
    ap = argparse.ArgumentParser(
        description="Create Gmail drafts from a Google Doc + a Google Sheet of recipients.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--doc", required=True, help="Google Doc URL or id (the message)")
    ap.add_argument("--sheet", required=True,
                    help="Recipients: Google Sheet or uploaded .xlsx, URL or id")
    ap.add_argument("--tab", help="Sheet tab name (default: first tab)")
    ap.add_argument("--section",
                    help="Use only this section of a multi-draft Doc, e.g. 'Email 1'")
    ap.add_argument("--filter", action="append", default=[], metavar="COL=VAL",
                    help="Only rows where COL equals VAL, e.g. --filter tier=1 "
                         "(repeatable; VAL may be a comma-separated list)")
    ap.add_argument("--exclude", action="append", default=[], metavar="COL=VAL",
                    help="Drop rows where COL equals VAL, e.g. "
                         "--exclude company='Jivi Health' (repeatable)")
    ap.add_argument("--header-row", type=int,
                    help="1-indexed header row (default: auto-detected)")
    ap.add_argument("--map", action="append", default=[], metavar="PLACEHOLDER=COL",
                    dest="maps",
                    help="Fill a placeholder from a different column, e.g. "
                         "--map name=first_name (repeatable)")
    ap.add_argument("--subject", help="Subject line, overrides the Doc's 'Subject:' line")
    ap.add_argument("--cc", action="append", default=[], help="CC on every draft (repeatable)")
    ap.add_argument("--bcc", action="append", default=[], help="BCC on every draft (repeatable)")
    ap.add_argument("--attach", action="append", default=[], help="Attach a file to every draft (repeatable)")
    ap.add_argument("--limit", type=int, help="Only process the first N eligible rows")
    ap.add_argument("--dry-run", action="store_true", help="Preview; create nothing")
    ap.add_argument("--force", action="store_true", help="Re-draft rows that already have a status")
    ap.add_argument("--no-mark", action="store_true", help="Don't write status back to the sheet")
    ap.add_argument(
        "--port",
        type=int,
        default=0,
        help=(
            "Fixed localhost port for the one-time OAuth callback. Only needed if "
            "credentials.json is a Web-application client, which requires an exact "
            "registered redirect URI (a Desktop client accepts any port)."
        ),
    )
    ap.add_argument(
        "--allow-missing",
        action="store_true",
        help="Draft even if a {{placeholder}} has no value (default: skip the row)",
    )
    args = ap.parse_args()

    doc_id = extract_id(args.doc, "Doc")
    sheet_id = extract_id(args.sheet, "Sheet")

    print("\n" + "=" * 62)
    print("  BATCH GMAIL DRAFTS — Doc + Sheet → drafts (never sends)")
    print("=" * 62 + "\n")

    filters = parse_filters(args.filter)
    excludes = parse_filters(args.exclude)
    aliases = parse_maps(args.maps)

    gmail, docs, sheets, drive, sender = authenticate(port=args.port)
    doc_subject, body_text_tpl, body_html_tpl = load_doc(docs, doc_id, args.section)
    headers, rows, tab_name, writable = load_sheet(
        sheets, drive, sheet_id, args.tab, args.header_row
    )

    subject_tpl = args.subject or doc_subject
    if not subject_tpl and "subject" not in headers:
        sys.exit(
            "❌ No subject. Add a 'Subject: ...' first line to the Doc, a 'subject' "
            "column to the Sheet, or pass --subject."
        )

    print(f"  📋 {len(rows)} data row(s); columns: {', '.join(h for h in headers if h)}")

    for label, specs in (("--filter", filters), ("--exclude", excludes)):
        for col, _ in specs:
            if col not in headers:
                sys.exit(f"❌ {label} column '{col}' is not in this tab.")

    if filters:
        before = len(rows)
        rows = [r for r in rows if passes_filters(r, filters)]
        shown = ", ".join(f"{c}={'/'.join(sorted(v))}" for c, v in filters)
        print(f"  🔎 Filter {shown}: {len(rows)} of {before} row(s) match")

    if excludes:
        dropped = [r for r in rows if is_excluded(r, excludes)]
        rows = [r for r in rows if not is_excluded(r, excludes)]
        shown = ", ".join(f"{c}={'/'.join(sorted(v))}" for c, v in excludes)
        who = ", ".join(r.get("name") or pick_email(r) for r in dropped) or "nobody"
        print(f"  🚫 Exclude {shown}: dropped {len(dropped)} ({who})")

    has_status = "status" in headers and writable
    has_draft_id = "draft_id" in headers and writable
    if not writable and not args.dry_run:
        print("  ℹ️  .xlsx source — status can't be written back; see drafts_log.csv")
    print()

    if args.dry_run:
        print("  🔍 DRY RUN — no drafts will be created\n")

    created = 0
    skipped = 0
    updates = []
    log_rows = []
    seen_emails = set()

    for row in rows:
        row = enrich(row)
        n = row["_row_number"]
        email = pick_email(row)
        label = row.get("first_name") or row.get("name") or email or f"row {n}"

        if not email:
            print(f"  [row {n}] ⏭️  no email address — skipped")
            skipped += 1
            continue
        if not EMAIL_RE.match(email):
            print(f"  [row {n}] ⏭️  invalid email '{email}' — skipped")
            skipped += 1
            continue
        if "status" in headers and row.get("status") and not args.force:
            print(f"  [row {n}] ⏭️  already done ({row['status']}) — skipped")
            skipped += 1
            continue
        if email.lower() in seen_emails:
            print(f"  [row {n}] ⏭️  duplicate of an earlier row ({email}) — skipped")
            skipped += 1
            continue

        subject_src = row.get("subject") or subject_tpl
        subject, miss_s = merge(subject_src, row, False, aliases)
        text, miss_t = merge(body_text_tpl, row, False, aliases)
        html_body, miss_h = merge(body_html_tpl, row, True, aliases)

        missing = sorted(set(miss_s + miss_t + miss_h))
        if missing and not args.allow_missing:
            print(f"  [row {n}] ⏭️  {label}: no value for {{{{{'}}, {{'.join(missing)}}}}} — skipped")
            skipped += 1
            continue
        if missing:
            print(f"  [row {n}] ⚠️  {label}: unfilled {', '.join(missing)}")

        cc = list(args.cc) + split_addresses(row.get("cc", ""))
        bcc = list(args.bcc) + split_addresses(row.get("bcc", ""))
        attachments = list(args.attach) + [
            a for a in re.split(r"[;,]", row.get("attachments", "")) if a.strip()
        ]

        seen_emails.add(email.lower())

        if args.dry_run:
            print(f"  [row {n}] 🔍 To: {email}")
            print(f"            Subject: {subject}")
            if cc:
                print(f"            Cc: {', '.join(cc)}")
            if bcc:
                print(f"            Bcc: {', '.join(bcc)}")
            if attachments:
                print(f"            Attach: {', '.join(attachments)}")
            preview = [ln for ln in text.split("\n")][:6]
            for ln in preview:
                print(f"            | {ln}")
            print(f"            | ... ({len(text.split(chr(10)))} lines total)\n")
            created += 1
        else:
            try:
                msg = build_message(
                    sender, email, subject, text, html_body, cc, bcc, attachments
                )
                draft_id = create_draft(gmail, msg)
                created += 1
                print(f"  [row {n}] 📧 {label} → {email}")
                if has_status and not args.no_mark:
                    updates.append((n, "status", f"Draft created {date.today().isoformat()}"))
                if has_draft_id and not args.no_mark:
                    updates.append((n, "draft_id", draft_id))
                log_rows.append({
                    "date": date.today().isoformat(),
                    "row": n,
                    "email": email,
                    "subject": subject,
                    "draft_id": draft_id,
                    "sheet_id": sheet_id,
                    "doc_id": doc_id,
                })
            except Exception as e:
                print(f"  [row {n}] ❌ {label}: {e}")
                skipped += 1

        if args.limit and created >= args.limit:
            print(f"\n  ⏹  Stopping at --limit {args.limit}")
            break

    if updates and not args.dry_run:
        try:
            write_back(sheets, sheet_id, tab_name, headers, updates)
        except Exception as e:
            print(f"  ⚠️  Sheet write-back failed (drafts are fine): {e}")

    if log_rows:
        new_file = not os.path.exists(LOG_PATH)
        with open(LOG_PATH, "a", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(log_rows[0].keys()))
            if new_file:
                w.writeheader()
            w.writerows(log_rows)
        print(f"  💾 Logged {len(log_rows)} draft(s) to {LOG_PATH}")

    print("\n" + "=" * 62)
    verb = "would be created" if args.dry_run else "created"
    print(f"  ✅ {created} draft(s) {verb}   ⏭️  {skipped} skipped")
    if not args.dry_run and created:
        print("  📬 Review them in Gmail → Drafts. Nothing was sent.")
    print("=" * 62 + "\n")


if __name__ == "__main__":
    main()
